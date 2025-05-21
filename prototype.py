import os
import requests
import json
import time
import base64
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def get_all_org_repos():
    """Get all repositories for the organization"""
    github_token = os.environ.get('GITHUB_TOKEN')
    org_name = os.environ.get('ORG_NAME')
    
    if not github_token or not org_name:
        print(f"Error: Required environment variables not found.")
        print(f"GITHUB_TOKEN: {'Found' if github_token else 'Missing'}")
        print(f"ORG_NAME: {'Found' if org_name else 'Missing'}")
        return None
    
    print(f"Fetching repositories for organization: {org_name}")
    
    headers = {
        'Authorization': f'token {github_token}',
        'Accept': 'application/vnd.github.v3+json'
    }
    
    try:
        all_repos = []
        page = 1
        per_page = 100
        
        while True:
            print(f"Fetching page {page} of repositories...")
            repos_endpoint = f'https://api.github.com/orgs/{org_name}/repos?page={page}&per_page={per_page}&type=all'
            repos_response = requests.get(repos_endpoint, headers=headers)
            
            # Handle rate limiting
            if handle_rate_limiting(repos_response):
                continue
            
            if repos_response.status_code != 200:
                print(f"Error: Could not retrieve repositories. Status code: {repos_response.status_code}")
                print(f"Response: {repos_response.text}")
                return None
            
            repos_page = repos_response.json()
            if not repos_page:
                break
                
            all_repos.extend(repos_page)
            print(f"Retrieved {len(repos_page)} repositories on page {page}. Total so far: {len(all_repos)}")
            
            if 'Link' in repos_response.headers and 'rel="next"' in repos_response.headers['Link']:
                page += 1
                time.sleep(0.5)
            else:
                break
        
        print(f"\nFound {len(all_repos)} repositories for organization '{org_name}'")
        return all_repos
        
    except requests.exceptions.RequestException as e:
        print(f"Error: {e}")
        return None

def handle_rate_limiting(response):
    """Handle GitHub API rate limiting"""
    if 'X-RateLimit-Remaining' in response.headers:
        remaining = int(response.headers['X-RateLimit-Remaining'])
        
        if remaining == 0 or (response.status_code == 403 and 'rate limit exceeded' in response.text.lower()):
            reset_time = int(response.headers.get('X-RateLimit-Reset', time.time() + 3600))
            wait_time = max(reset_time - time.time(), 0) + 1
            print(f"API rate limit reached. Waiting for {wait_time:.2f} seconds...")
            time.sleep(wait_time)
            return True
    
    return False

def find_python_project_files(repos):
    """Find repositories with pyproject.toml and requirements.txt files"""
    github_token = os.environ.get('GITHUB_TOKEN')
    org_name = os.environ.get('ORG_NAME')
    
    headers = {
        'Authorization': f'token {github_token}',
        'Accept': 'application/vnd.github.v3+json'
    }
    
    python_repos = []
    
    print("\nSearching for Python project files in repositories...")
    print("Looking for pyproject.toml and requirements.txt files")
    
    for index, repo in enumerate(repos, 1):
        repo_name = repo.get('name')
        print(f"[{index}/{len(repos)}] Checking repository: {repo_name}")
        
        contents_endpoint = f'https://api.github.com/repos/{org_name}/{repo_name}/contents'
        contents_response = requests.get(contents_endpoint, headers=headers)
        
        if handle_rate_limiting(contents_response):
            # If we hit a rate limit, retry this repository
            index -= 1
            continue
        
        if contents_response.status_code != 200:
            print(f"  Unable to access contents for {repo_name}. Status: {contents_response.status_code}")
            continue
        
        try:
            contents = contents_response.json()
            
            # Check if the response is a list (directory listing)
            if not isinstance(contents, list):
                print(f"  Unexpected response format for {repo_name}")
                continue
            
            # Look for Python project files
            has_pyproject = any(item.get('name') == 'pyproject.toml' for item in contents)
            has_requirements = any(item.get('name') == 'requirements.txt' for item in contents)
            has_setup_py = any(item.get('name') == 'setup.py' for item in contents)
            
            # Get content of requirements.txt if found
            requirements_content = None
            requirements_packages = []
            if has_requirements:
                for item in contents:
                    if item.get('name') == 'requirements.txt':
                        req_response = requests.get(item.get('url'), headers=headers)
                        if req_response.status_code == 200:
                            req_data = req_response.json()
                            if req_data.get('encoding') == 'base64':
                                requirements_content = base64.b64decode(req_data.get('content')).decode('utf-8')
                                # Parse requirements file to extract packages
                                for line in requirements_content.splitlines():
                                    line = line.strip()
                                    if line and not line.startswith('#'):
                                        package_info = parse_requirement_line(line)
                                        if package_info:
                                            requirements_packages.append(package_info)
            
            # Get content of pyproject.toml if found
            pyproject_content = None
            pyproject_packages = []
            if has_pyproject:
                for item in contents:
                    if item.get('name') == 'pyproject.toml':
                        pyproj_response = requests.get(item.get('url'), headers=headers)
                        if pyproj_response.status_code == 200:
                            pyproj_data = pyproj_response.json()
                            if pyproj_data.get('encoding') == 'base64':
                                pyproject_content = base64.b64decode(pyproj_data.get('content')).decode('utf-8')
                                # We'll parse these later as they require more complex toml parsing
            
            # Only add repositories that have Python project files
            if has_pyproject or has_requirements or has_setup_py:
                python_repos.append({
                    'name': repo_name,
                    'url': repo.get('html_url'),
                    'description': repo.get('description'),
                    'has_pyproject': has_pyproject,
                    'has_requirements': has_requirements,
                    'has_setup_py': has_setup_py,
                    'requirements_content': requirements_content,
                    'requirements_packages': requirements_packages,
                    'pyproject_content': pyproject_content,
                    'pyproject_packages': pyproject_packages,
                    'language': repo.get('language'),
                    'last_updated': repo.get('updated_at'),
                    'created_at': repo.get('created_at'),
                    'stars': repo.get('stargazers_count'),
                    'forks': repo.get('forks_count'),
                    'default_branch': repo.get('default_branch')
                })
                
                print(f"  Found Python project: {repo_name}")
                if has_requirements:
                    print(f"  - Has requirements.txt with {len(requirements_packages)} packages")
                if has_pyproject:
                    print(f"  - Has pyproject.toml")
                if has_setup_py:
                    print(f"  - Has setup.py")
            
        except Exception as e:
            print(f"  Error processing repository {repo_name}: {e}")
    
    return python_repos

def parse_requirement_line(line):
    """Parse a line from requirements.txt to extract package name and version"""
    # Remove comments
    line = line.split('#')[0].strip()
    if not line:
        return None
    
    # Handle git/url requirements
    if line.startswith('git+') or line.startswith('http'):
        # Extract the package name from git/url if possible
        parts = line.split('#egg=')
        if len(parts) > 1:
            return {'name': parts[1].strip(), 'version': 'git', 'raw': line}
        return {'name': line, 'version': 'url', 'raw': line}
    
    # Handle standard requirements with versions
    for operator in ['==', '>=', '<=', '>', '<', '~=', '!=']:
        if operator in line:
            parts = line.split(operator, 1)
            return {'name': parts[0].strip(), 'version': f"{operator}{parts[1].strip()}", 'raw': line}
    
    # Just a package name without version
    return {'name': line.strip(), 'version': 'latest', 'raw': line}

def create_excel_report(python_repos):
    """Create a well-formatted Excel report with repositories and their packages"""
    print("\nCreating detailed Excel report...")
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Python Repositories"
    
    # Set column headers
    headers = [
        "Repository", "Description", "Language", "Last Updated", 
        "Has pyproject.toml", "Has requirements.txt", "Has setup.py", 
        "Package Count", "Packages List", "URL"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        # Add border
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    # Add data rows
    for row_num, repo in enumerate(python_repos, 2):
        # Extract packages from requirements.txt
        packages = repo['requirements_packages']
        package_list = ", ".join([f"{p['name']}{p['version'] if p['version'] != 'latest' else ''}" for p in packages])
        
        # Add repository data
        ws.cell(row=row_num, column=1).value = repo['name']
        ws.cell(row=row_num, column=2).value = repo['description'] or "No description"
        ws.cell(row=row_num, column=3).value = repo['language'] or "Unknown"
        ws.cell(row=row_num, column=4).value = repo['last_updated']
        ws.cell(row=row_num, column=5).value = "Yes" if repo['has_pyproject'] else "No"
        ws.cell(row=row_num, column=6).value = "Yes" if repo['has_requirements'] else "No"
        ws.cell(row=row_num, column=7).value = "Yes" if repo['has_setup_py'] else "No"
        ws.cell(row=row_num, column=8).value = len(packages)
        ws.cell(row=row_num, column=9).value = package_list
        ws.cell(row=row_num, column=10).value = repo['url']
        
        # Add hyperlink to repository URL
        ws.cell(row=row_num, column=10).hyperlink = repo['url']
        ws.cell(row=row_num, column=10).font = Font(color="0563C1", underline="single")
        
        # Apply formatting
        for col_num in range(1, 11):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if col_num == 9:  # Packages list column
                cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Create second sheet with detailed package analysis
    create_package_matrix_sheet(wb, python_repos)
    
    # Set column widths
    column_widths = {
        1: 25,  # Repository name
        2: 30,  # Description
        3: 12,  # Language
        4: 20,  # Last Updated
        5: 15,  # Has pyproject.toml
        6: 15,  # Has requirements.txt
        7: 15,  # Has setup.py
        8: 12,  # Package Count
        9: 60,  # Packages List
        10: 45,  # URL
    }
    
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Save the workbook
    excel_filename = "python_repositories_report.xlsx"
    wb.save(excel_filename)
    print(f"Excel report saved as {excel_filename}")

def create_package_matrix_sheet(workbook, python_repos):
    """Create a matrix sheet showing repositories vs packages"""
    # Get a list of all unique packages across all repositories
    all_packages = set()
    for repo in python_repos:
        for pkg in repo['requirements_packages']:
            all_packages.add(pkg['name'])
    
    # Sort packages alphabetically
    all_packages = sorted(list(all_packages))
    
    # Create the matrix sheet
    ws = workbook.create_sheet(title="Packages Matrix")
    
    # Add header row with package names
    ws.cell(row=1, column=1).value = "Repository"
    for col_num, package in enumerate(all_packages, 2):
        ws.cell(row=1, column=col_num).value = package
        ws.cell(row=1, column=col_num).font = Font(bold=True)
        ws.cell(row=1, column=col_num).alignment = Alignment(textRotation=90, horizontal='center')
    
    # Add repository rows
    for row_num, repo in enumerate(python_repos, 2):
        # Repository name in first column
        ws.cell(row=row_num, column=1).value = repo['name']
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        
        # Get this repo's packages
        repo_packages = {pkg['name']: pkg['version'] for pkg in repo['requirements_packages']}
        
        # Fill in the matrix
        for col_num, package in enumerate(all_packages, 2):
            if package in repo_packages:
                version = repo_packages[package]
                cell = ws.cell(row=row_num, column=col_num)
                
                # Display version if available
                if version != 'latest':
                    cell.value = version
                else:
                    cell.value = "✓"
                
                # Highlight cell
                cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                cell.alignment = Alignment(horizontal='center')
    
    # Set column width for repository names
    ws.column_dimensions['A'].width = 25
    
    # Set width for package columns
    for col_num in range(2, len(all_packages) + 2):
        ws.column_dimensions[get_column_letter(col_num)].width = 10
    
    # Freeze first column and header row
    ws.freeze_panes = "B2"

def create_summary_sheet(workbook, python_repos):
    """Create a summary sheet with key statistics"""
    ws = workbook.create_sheet(title="Summary")
    
    # Calculate statistics
    total_repos = len(python_repos)
    repos_with_requirements = sum(1 for repo in python_repos if repo['has_requirements'])
    repos_with_pyproject = sum(1 for repo in python_repos if repo['has_pyproject'])
    repos_with_setup_py = sum(1 for repo in python_repos if repo['has_setup_py'])
    
    # Count total unique packages
    all_packages = set()
    for repo in python_repos:
        for pkg in repo['requirements_packages']:
            all_packages.add(pkg['name'])
    
    # Most common packages
    package_counts = {}
    for repo in python_repos:
        for pkg in repo['requirements_packages']:
            package_counts[pkg['name']] = package_counts.get(pkg['name'], 0) + 1
    
    top_packages = sorted(package_counts.items(), key=lambda x: x[1], reverse=True)[:10]
    
    # Add summary data
    row = 1
    ws.cell(row=row, column=1).value = "Python Repositories Summary"
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    
    row += 2
    ws.cell(row=row, column=1).value = "Total Python Repositories:"
    ws.cell(row=row, column=2).value = total_repos
    
    row += 1
    ws.cell(row=row, column=1).value = "Repos with requirements.txt:"
    ws.cell(row=row, column=2).value = repos_with_requirements
    ws.cell(row=row, column=3).value = f"{repos_with_requirements/total_repos*100:.1f}%"
    
    row += 1
    ws.cell(row=row, column=1).value = "Repos with pyproject.toml:"
    ws.cell(row=row, column=2).value = repos_with_pyproject
    ws.cell(row=row, column=3).value = f"{repos_with_pyproject/total_repos*100:.1f}%"
    
    row += 1
    ws.cell(row=row, column=1).value = "Repos with setup.py:"
    ws.cell(row=row, column=2).value = repos_with_setup_py
    ws.cell(row=row, column=3).value = f"{repos_with_setup_py/total_repos*100:.1f}%"
    
    row += 1
    ws.cell(row=row, column=1).value = "Total unique packages:"
    ws.cell(row=row, column=2).value = len(all_packages)
    
    # Add top packages section
    row += 2
    ws.cell(row=row, column=1).value = "Top 10 Most Common Packages"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    
    row += 1
    ws.cell(row=row, column=1).value = "Package"
    ws.cell(row=row, column=2).value = "Count"
    ws.cell(row=row, column=3).value = "Percentage"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3).font = Font(bold=True)
    
    for package, count in top_packages:
        row += 1
        ws.cell(row=row, column=1).value = package
        ws.cell(row=row, column=2).value = count
        ws.cell(row=row, column=3).value = f"{count/total_repos*100:.1f}%"
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

def main():
    """Main function to execute the script"""
    print("Starting GitHub repository scanner...")
    
    # Get all repositories in the organization
    repos = get_all_org_repos()
    if not repos:
        print("No repositories found or an error occurred.")
        return
    
    # Find repositories with Python project files
    python_repos = find_python_project_files(repos)
    
    if not python_repos:
        print("No Python projects found in the organization repositories.")
        return
    
    print(f"\nFound {len(python_repos)} repositories with Python project files.")
    
    # Create Excel report
    create_excel_report(python_repos)
    
    print("\nScript completed successfully.")

if __name__ == "__main__":
    main()