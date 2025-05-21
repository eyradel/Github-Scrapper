import os
import requests
import json
import time
import base64
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

def get_all_org_repos(limit=None):
    """Get repositories for the organization with optional limit"""
    github_token = os.environ.get('GITHUB_TOKEN')
    org_name = os.environ.get('ORG_NAME')
    
    if not github_token or not org_name:
        print(f"Error: Required environment variables not found.")
        print(f"GITHUB_TOKEN: {'Found' if github_token else 'Missing'}")
        print(f"ORG_NAME: {'Found' if org_name else 'Missing'}")
        return None
    
    print(f"Fetching repositories for organization: {org_name}" + (f" (limited to {limit})" if limit else ""))
    
    # First verify token permissions
    headers = {
        'Authorization': f'token {github_token}',
        'Accept': 'application/vnd.github.v3+json'
    }
    
    try:
        # Check token permissions
        user_response = requests.get('https://api.github.com/user', headers=headers)
        if user_response.status_code != 200:
            print(f"Error: Unable to authenticate with token. Status: {user_response.status_code}")
            return None
            
        user_data = user_response.json()
        print(f"Authenticated as: {user_data.get('login')}")
        
        # Check token scopes
        if 'X-OAuth-Scopes' in user_response.headers:
            scopes = user_response.headers['X-OAuth-Scopes']
            print(f"Token scopes: {scopes}")
            if 'repo' not in scopes:
                print("WARNING: Your token may not have full 'repo' scope, which is needed to see private repositories.")
            if 'read:org' not in scopes and 'admin:org' not in scopes:
                print("WARNING: Your token may not have organization reading permissions.")
        
        # Fetch repositories with effective pagination
        all_repos = []
        page = 1
        per_page = 100  # Maximum allowed by GitHub API
        
        # Use both member endpoint and regular endpoint to get maximum coverage
        endpoints = [
            f'https://api.github.com/orgs/{org_name}/repos',  # Standard API
            f'https://api.github.com/user/repos'  # User's repos (includes private ones user has access to)
        ]
        
        for endpoint_base in endpoints:
            print(f"\nUsing endpoint: {endpoint_base}")
            page = 1
            
            # Check if we've reached the limit
            if limit and len(all_repos) >= limit:
                print(f"Reached specified limit of {limit} repositories. Stopping search.")
                break
                
            while True:
                # Carefully construct parameters to get ALL repos
                params = {
                    'page': page,
                    'per_page': per_page,
                    'type': 'all',  # Include all types (public, private, etc.)
                    'sort': 'full_name',  # Consistent sorting to avoid duplicates
                    'direction': 'asc'
                }
                
                # For user repos endpoint, filter by org
                if 'user/repos' in endpoint_base:
                    params['affiliation'] = 'organization_member'
                
                print(f"Fetching page {page} with {per_page} repositories per page...")
                response = requests.get(endpoint_base, headers=headers, params=params)
                
                # Handle rate limiting
                if handle_rate_limiting(response):
                    continue
                
                if response.status_code != 200:
                    print(f"Error: Could not retrieve repositories. Status code: {response.status_code}")
                    print(f"Response: {response.text}")
                    break
                
                repos_page = response.json()
                
                if not repos_page:
                    print(f"No repositories found on page {page}")
                    break
                
                # Filter to only include repos from our organization
                if 'user/repos' in endpoint_base:
                    org_repos = [r for r in repos_page if r.get('owner', {}).get('login') == org_name]
                    print(f"Found {len(org_repos)} repositories for org '{org_name}' from {len(repos_page)} total")
                    repos_page = org_repos
                
                # Check for duplicates before adding
                new_repos = []
                existing_names = {r['name'] for r in all_repos}
                for repo in repos_page:
                    if repo['name'] not in existing_names:
                        new_repos.append(repo)
                        existing_names.add(repo['name'])
                
                all_repos.extend(new_repos)
                print(f"Added {len(new_repos)} new repositories. Total unique repos so far: {len(all_repos)}")
                
                # Check if we've reached the limit after adding new repos
                if limit and len(all_repos) >= limit:
                    print(f"Reached specified limit of {limit} repositories. Stopping search.")
                    break
                
                # Check for more pages
                if 'Link' in response.headers:
                    links = response.headers['Link']
                    if 'rel="next"' not in links:
                        print("No more pages.")
                        break
                elif len(repos_page) < per_page:
                    print("Received fewer repositories than requested. This is the last page.")
                    break
                
                page += 1
                time.sleep(0.5)  # Be nice to the API
        
        # Apply limit if specified
        if limit and len(all_repos) > limit:
            all_repos = all_repos[:limit]
            print(f"Trimmed to first {limit} repositories as requested.")
        
        print(f"\nFound {len(all_repos)} total unique repositories for organization '{org_name}'")
        
        # Double check against expected number
        if limit is None and len(all_repos) < 183:
            print(f"WARNING: Found {len(all_repos)} repositories, but expected 183 based on the GitHub UI.")
            print("This might be due to permission issues or API limitations.")
        
        return all_repos
        
    except requests.exceptions.RequestException as e:
        print(f"Error: {e}")
        return None

def handle_rate_limiting(response):
    """Handle GitHub API rate limiting"""
    if 'X-RateLimit-Remaining' in response.headers:
        remaining = int(response.headers['X-RateLimit-Remaining'])
        
        if remaining < 10:
            print(f"Warning: Only {remaining} API requests remaining.")
        
        if remaining == 0 or (response.status_code == 403 and 'rate limit exceeded' in response.text.lower()):
            reset_time = int(response.headers.get('X-RateLimit-Reset', time.time() + 3600))
            wait_time = max(reset_time - time.time(), 0) + 1
            print(f"API rate limit reached. Waiting for {wait_time:.2f} seconds...")
            time.sleep(wait_time)
            return True
    
    return False

def get_repository_branches(repo_full_name):
    """Get all branches for a specific repository"""
    github_token = os.environ.get('GITHUB_TOKEN')
    
    headers = {
        'Authorization': f'token {github_token}',
        'Accept': 'application/vnd.github.v3+json'
    }
    
    all_branches = []
    page = 1
    per_page = 100
    
    while True:
        branches_endpoint = f'https://api.github.com/repos/{repo_full_name}/branches?page={page}&per_page={per_page}'
        branches_response = requests.get(branches_endpoint, headers=headers)
        
        if handle_rate_limiting(branches_response):
            continue
            
        if branches_response.status_code != 200:
            print(f"  Warning: Could not retrieve branches for {repo_full_name}. Status: {branches_response.status_code}")
            break
            
        branches_page = branches_response.json()
        if not branches_page:
            break
            
        all_branches.extend(branches_page)
        
        if len(branches_page) < per_page:
            break
            
        page += 1
        time.sleep(0.5)
    
    return all_branches

def find_file_in_branch(repo_full_name, branch_name, file_path):
    """Check if a specific file exists in a branch and return its content if found"""
    github_token = os.environ.get('GITHUB_TOKEN')
    
    headers = {
        'Authorization': f'token {github_token}',
        'Accept': 'application/vnd.github.v3+json'
    }
    
    file_endpoint = f'https://api.github.com/repos/{repo_full_name}/contents/{file_path}?ref={branch_name}'
    file_response = requests.get(file_endpoint, headers=headers)
    
    if handle_rate_limiting(file_response):
        return None
        
    if file_response.status_code != 200:
        return None
        
    try:
        file_data = file_response.json()
        if file_data.get('encoding') == 'base64':
            content = base64.b64decode(file_data.get('content')).decode('utf-8', errors='replace')
            return content
    except Exception:
        pass
        
    return None

def get_branch_tree(repo_full_name, branch_name):
    """Get the tree of files in a branch with robust error handling"""
    github_token = os.environ.get('GITHUB_TOKEN')
    
    headers = {
        'Authorization': f'token {github_token}',
        'Accept': 'application/vnd.github.v3+json'
    }
    
    # First, get the branch reference to get the SHA
    ref_endpoint = f'https://api.github.com/repos/{repo_full_name}/branches/{branch_name}'
    
    # Add retry logic for API requests
    max_retries = 3
    for attempt in range(max_retries):
        try:
            ref_response = requests.get(ref_endpoint, headers=headers, timeout=30)
            
            if handle_rate_limiting(ref_response):
                continue
                
            if ref_response.status_code != 200:
                print(f"  Warning: Could not get branch reference for {repo_full_name}:{branch_name}")
                return []
            
            branch_data = ref_response.json()
            commit_sha = branch_data.get('commit', {}).get('sha')
            
            if not commit_sha:
                print(f"  Warning: Could not get commit SHA for {repo_full_name}:{branch_name}")
                return []
            
            # Now get the tree recursively
            tree_endpoint = f'https://api.github.com/repos/{repo_full_name}/git/trees/{commit_sha}?recursive=1'
            
            try:
                tree_response = requests.get(tree_endpoint, headers=headers, timeout=60)  # Longer timeout
                
                if handle_rate_limiting(tree_response):
                    continue
                
                if tree_response.status_code != 200:
                    print(f"  Warning: Could not get tree for {repo_full_name}:{branch_name}")
                    return get_python_files_alternative(repo_full_name, branch_name, headers)
                
                tree_data = tree_response.json()
                
                # Check if tree is truncated
                if tree_data.get('truncated', False):
                    print(f"  Tree is truncated for {repo_full_name}:{branch_name}. Using alternative approach...")
                    return get_python_files_alternative(repo_full_name, branch_name, headers)
                
                tree_items = tree_data.get('tree', [])
                
                # Filter for Python files
                python_files = [item for item in tree_items if item.get('type') == 'blob' and item.get('path', '').endswith('.py')]
                
                return python_files
                
            except (requests.exceptions.RequestException, requests.exceptions.ChunkedEncodingError) as e:
                # If we fail to get the tree, try alternative approach
                print(f"  Error fetching tree: {e}. Trying alternative approach...")
                return get_python_files_alternative(repo_full_name, branch_name, headers)
                
        except (requests.exceptions.RequestException, json.JSONDecodeError) as e:
            wait_time = 2 ** attempt
            print(f"  Attempt {attempt+1}/{max_retries} failed: {e}. Retrying in {wait_time} seconds...")
            time.sleep(wait_time)
    
    # If all retries fail, try alternative approach
    print(f"  All attempts failed for {repo_full_name}:{branch_name}. Using alternative approach...")
    return get_python_files_alternative(repo_full_name, branch_name, headers)

def get_python_files_alternative(repo_full_name, branch_name, headers):
    """Alternative approach to find Python files when the tree API fails"""
    print(f"  Using alternative approach to find Python files in {repo_full_name}:{branch_name}")
    
    # Get the root directory contents
    root_endpoint = f'https://api.github.com/repos/{repo_full_name}/contents?ref={branch_name}'
    
    try:
        root_response = requests.get(root_endpoint, headers=headers, timeout=30)
        
        if handle_rate_limiting(root_response):
            return []
            
        if root_response.status_code != 200:
            print(f"  Could not get root directory for {repo_full_name}:{branch_name}")
            return []
        
        python_files = []
        root_contents = root_response.json()
        
        # Process root contents
        for item in root_contents:
            if item.get('type') == 'file' and item.get('name', '').endswith('.py'):
                # Add Python files from root
                python_files.append({
                    'path': item.get('path'),
                    'type': 'blob',
                    'url': item.get('url')
                })
            elif item.get('type') == 'dir':
                # Add files from important directories
                important_dirs = ['src', 'app', 'lib', 'core', 'models', 'utils']
                if item.get('name') in important_dirs:
                    dir_files = get_directory_python_files(repo_full_name, branch_name, item.get('path'), headers)
                    python_files.extend(dir_files)
        
        # If we found too few Python files, try checking some standard directories
        if len(python_files) < 5:
            for dir_name in ['src', 'app', 'lib']:
                dir_files = get_directory_python_files(repo_full_name, branch_name, dir_name, headers)
                python_files.extend(dir_files)
        
        return python_files
        
    except (requests.exceptions.RequestException, json.JSONDecodeError) as e:
        print(f"  Error in alternative approach: {e}")
        return []

def get_directory_python_files(repo_full_name, branch_name, dir_path, headers):
    """Get Python files from a specific directory"""
    dir_endpoint = f'https://api.github.com/repos/{repo_full_name}/contents/{dir_path}?ref={branch_name}'
    
    try:
        dir_response = requests.get(dir_endpoint, headers=headers, timeout=30)
        
        if dir_response.status_code != 200:
            return []
        
        python_files = []
        dir_contents = dir_response.json()
        
        # Process directory contents
        for item in dir_contents:
            if item.get('type') == 'file' and item.get('name', '').endswith('.py'):
                python_files.append({
                    'path': item.get('path'),
                    'type': 'blob',
                    'url': item.get('url')
                })
        
        return python_files
        
    except (requests.exceptions.RequestException, json.JSONDecodeError):
        # Silently fail for subdirectories
        return []

def extract_imports_from_python_content(content):
    """Extract import statements from Python file content"""
    imports = set()
    
    # Regular expressions for different types of imports
    import_patterns = [
        r'^\s*import\s+(\w+(?:\s*,\s*\w+)*)',                    # import module, module2
        r'^\s*from\s+(\w+(?:\.\w+)*)\s+import\s+',               # from module import
        r'^\s*from\s+(\w+(?:\.\w+)*)\s+import\s+\(',             # from module import (
        r'^\s*import\s+(\w+(?:\.\w+)*)\s+as\s+\w+',              # import module as alias
        r'^\s*from\s+(\w+(?:\.\w+)*)\s+import\s+\w+\s+as\s+\w+'  # from module import name as alias
    ]
    
    lines = content.split('\n')
    for line in lines:
        # Skip comments
        if line.strip().startswith('#'):
            continue
        
        for pattern in import_patterns:
            matches = re.findall(pattern, line, re.MULTILINE)
            for match in matches:
                # Handle multiple imports in one line (e.g., import os, sys)
                for module in match.split(','):
                    # Get the top-level package name (before any dots)
                    top_level = module.strip().split('.')[0]
                    if top_level:
                        imports.add(top_level)
    
    # Filter out standard library modules (this is a simple approach and not comprehensive)
    std_lib = {
        'abc', 'argparse', 'asyncio', 'collections', 'concurrent', 'contextlib', 'copy', 
        'csv', 'datetime', 'decimal', 'email', 'enum', 'functools', 'glob', 'hashlib', 
        'http', 'importlib', 'inspect', 'io', 'itertools', 'json', 'logging', 'math', 
        'multiprocessing', 'operator', 'os', 'pathlib', 'pickle', 'random', 're', 
        'shutil', 'signal', 'socket', 'sqlite3', 'statistics', 'string', 'subprocess', 
        'sys', 'tempfile', 'threading', 'time', 'traceback', 'typing', 'uuid', 'warnings', 
        'weakref', 'xml', 'zipfile'
    }
    
    # Return as a set of tuples with a flag indicating if it's a standard library
    return {(imp, imp in std_lib) for imp in imports}

def get_python_file_content(repo_full_name, branch_name, file_path):
    """Get the content of a Python file from a branch"""
    github_token = os.environ.get('GITHUB_TOKEN')
    
    headers = {
        'Authorization': f'token {github_token}',
        'Accept': 'application/vnd.github.v3+json'
    }
    
    content_endpoint = f'https://api.github.com/repos/{repo_full_name}/contents/{file_path}?ref={branch_name}'
    content_response = requests.get(content_endpoint, headers=headers)
    
    if handle_rate_limiting(content_response):
        return None
        
    if content_response.status_code != 200:
        return None
    
    content_data = content_response.json()
    if content_data.get('encoding') == 'base64':
        try:
            content = base64.b64decode(content_data.get('content')).decode('utf-8', errors='replace')
            return content
        except Exception as e:
            print(f"  Error decoding file {file_path}: {e}")
    
    return None

def analyze_python_files_in_branch(repo_full_name, branch_name):
    """Analyze all Python files in a branch to extract imported libraries"""
    print(f"  Analyzing Python files in branch: {branch_name}")
    
    # Get all Python files in the branch
    python_files = get_branch_tree(repo_full_name, branch_name)
    if not python_files:
        print(f"  No Python files found in branch {branch_name}")
        return []
    
    print(f"  Found {len(python_files)} Python files in branch {branch_name}")
    
    # Analyze a subset of files to avoid API rate limits (max 50 files per branch)
    max_files = 50
    if len(python_files) > max_files:
        print(f"  Limiting analysis to {max_files} Python files to avoid API rate limits")
        python_files = python_files[:max_files]
    
    # Process each Python file
    all_imports = set()
    processed_files = 0
    
    for python_file in python_files:
        file_path = python_file.get('path')
        if not file_path:
            continue
        
        print(f"    Analyzing file: {file_path}")
        
        # Get file content
        content = get_python_file_content(repo_full_name, branch_name, file_path)
        if not content:
            print(f"    Could not retrieve content for {file_path}")
            continue
        
        # Extract imports
        file_imports = extract_imports_from_python_content(content)
        all_imports.update(file_imports)
        
        processed_files += 1
        
        # Add a delay to avoid hitting API rate limits
        time.sleep(0.1)
    
    print(f"  Processed {processed_files} Python files")
    print(f"  Found {len(all_imports)} unique imports")
    
    return list(all_imports)

def find_python_project_files(repos):
    """Find repositories with Python files and analyze their imports"""
    github_token = os.environ.get('GITHUB_TOKEN')
    org_name = os.environ.get('ORG_NAME')
    
    headers = {
        'Authorization': f'token {github_token}',
        'Accept': 'application/vnd.github.v3+json'
    }
    
    python_repos = []
    
    print("\nSearching for Python files in repositories...")
    
    for index, repo in enumerate(repos, 1):
        repo_name = repo.get('name')
        repo_full_name = repo.get('full_name', f"{org_name}/{repo_name}")
        print(f"[{index}/{len(repos)}] Checking repository: {repo_name}")
        
        # Get all branches for this repository
        branches = get_repository_branches(repo_full_name)
        print(f"  Found {len(branches)} branches")
        
        repo_data = {
            'name': repo_name,
            'full_name': repo_full_name,
            'url': repo.get('html_url'),
            'description': repo.get('description'),
            'language': repo.get('language'),
            'last_updated': repo.get('updated_at'),
            'created_at': repo.get('created_at'),
            'stars': repo.get('stargazers_count'),
            'forks': repo.get('forks_count'),
            'default_branch': repo.get('default_branch'),
            'branches': [],
            'has_python_files': False
        }
        
        # Check each branch for Python project files
        for branch in branches:
            branch_name = branch.get('name')
            print(f"  Checking branch: {branch_name}")
            
            branch_data = {
                'name': branch_name,
                'is_default': branch_name == repo.get('default_branch'),
                'has_pyproject': False,
                'has_requirements': False,
                'has_setup_py': False,
                'requirements_content': None,
                'requirements_packages': [],
                'pyproject_content': None,
                'python_imports': [],
                'python_files_analyzed': 0
            }
            
            # Check for requirements.txt
            requirements_content = find_file_in_branch(repo_full_name, branch_name, 'requirements.txt')
            if requirements_content:
                branch_data['has_requirements'] = True
                branch_data['requirements_content'] = requirements_content
                
                # Parse requirements.txt to extract packages
                packages = []
                for line in requirements_content.splitlines():
                    line = line.strip()
                    if line and not line.startswith('#'):
                        package_info = parse_requirement_line(line)
                        if package_info:
                            packages.append(package_info)
                
                branch_data['requirements_packages'] = packages
                print(f"    Found requirements.txt with {len(packages)} packages")
            
            # Check for pyproject.toml
            pyproject_content = find_file_in_branch(repo_full_name, branch_name, 'pyproject.toml')
            if pyproject_content:
                branch_data['has_pyproject'] = True
                branch_data['pyproject_content'] = pyproject_content
                print(f"    Found pyproject.toml")
            
            # Check for setup.py
            setup_py_content = find_file_in_branch(repo_full_name, branch_name, 'setup.py')
            if setup_py_content:
                branch_data['has_setup_py'] = True
                print(f"    Found setup.py")
            
            # Analyze Python files for imports
            if repo.get('language') == 'Python' or branch_data['has_pyproject'] or branch_data['has_requirements'] or branch_data['has_setup_py']:
                python_imports = analyze_python_files_in_branch(repo_full_name, branch_name)
                branch_data['python_imports'] = python_imports
                branch_data['python_files_analyzed'] = True
                
                # Add branch data if Python files or project files were found
                if python_imports or branch_data['has_pyproject'] or branch_data['has_requirements'] or branch_data['has_setup_py']:
                    repo_data['branches'].append(branch_data)
                    repo_data['has_python_files'] = True
            else:
                # Skip analyzing Python files if the repository doesn't look like a Python project
                print(f"    Skipping Python file analysis for branch {branch_name} (not a Python project)")
        
        # Only include repositories that have Python files in at least one branch
        if repo_data['has_python_files']:
            python_repos.append(repo_data)
            print(f"  Found Python files in {len(repo_data['branches'])} branches")
    
    return python_repos

def parse_requirement_line(line):
    """Parse a line from requirements.txt to extract package name and version"""
    # Remove comments
    line = line.split('#')[0].strip()
    if not line:
        return None
    
    # Handle git/url requirements
    if line.startswith('git+') or line.startswith('http'):
        # Extract the package name from git/url if possible
        parts = line.split('#egg=')
        if len(parts) > 1:
            return {'name': parts[1].strip(), 'version': 'git', 'raw': line}
        return {'name': line, 'version': 'url', 'raw': line}
    
    # Handle standard requirements with versions
    for operator in ['==', '>=', '<=', '>', '<', '~=', '!=']:
        if operator in line:
            parts = line.split(operator, 1)
            return {'name': parts[0].strip(), 'version': f"{operator}{parts[1].strip()}", 'raw': line}
    
    # Just a package name without version
    return {'name': line.strip(), 'version': 'latest', 'raw': line}

def create_excel_report(python_repos):
    """Create a comprehensive Excel report with repositories, branches, and packages"""
    print("\nCreating detailed Excel report...")
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Repositories Overview"
    
    # Set column headers for the overview sheet
    headers = [
        "Repository", "Description", "Language", "Last Updated", 
        "Default Branch", "Total Branches", "Branches with Python Files",
        "Stars", "Forks", "URL"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        
        # Add border
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        cell.border = thin_border
    
    # Add repository overview data
    for row_num, repo in enumerate(python_repos, 2):
        total_branches = len(repo.get('branches', []))
        branches_with_python = len([b for b in repo.get('branches', []) if b.get('has_pyproject') or b.get('has_requirements') or b.get('has_setup_py') or b.get('python_imports')])
        
        ws.cell(row=row_num, column=1).value = repo['name']
        ws.cell(row=row_num, column=2).value = repo.get('description') or "No description"
        ws.cell(row=row_num, column=3).value = repo.get('language') or "Unknown"
        ws.cell(row=row_num, column=4).value = repo.get('last_updated')
        ws.cell(row=row_num, column=5).value = repo.get('default_branch')
        ws.cell(row=row_num, column=6).value = total_branches
        ws.cell(row=row_num, column=7).value = branches_with_python
        ws.cell(row=row_num, column=8).value = repo.get('stars', 0)
        ws.cell(row=row_num, column=9).value = repo.get('forks', 0)
        ws.cell(row=row_num, column=10).value = repo['url']
        
        # Add hyperlink to repository URL
        ws.cell(row=row_num, column=10).hyperlink = repo['url']
        ws.cell(row=row_num, column=10).font = Font(color="0563C1", underline="single")
        
        # Apply formatting
        for col_num in range(1, 11):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    # Set column widths
    column_widths = {
        1: 25,  # Repository name
        2: 40,  # Description
        3: 12,  # Language
        4: 20,  # Last Updated
        5: 15,  # Default Branch
        6: 12,  # Total Branches
        7: 15,  # Branches with Python Files
        8: 10,  # Stars
        9: 10,  # Forks
        10: 40,  # URL
    }
    
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Create a branches sheet with detailed information
    create_branches_sheet(wb, python_repos)
    
    # Create a packages matrix sheet
    create_packages_matrix(wb, python_repos)
    
    # Create an imports sheet
    create_imports_sheet(wb, python_repos)
    
    # Create a summary sheet
    create_summary_sheet(wb, python_repos)
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Save the workbook
    excel_filename = "python_repositories_analysis.xlsx"
    wb.save(excel_filename)
    print(f"Excel report saved as {excel_filename}")

def create_branches_sheet(workbook, python_repos):
    """Create a sheet with detailed branch information"""
    ws = workbook.create_sheet(title="Branch Details")
    
    # Set column headers
    headers = [
        "Repository", "Branch", "Is Default", "Has pyproject.toml", 
        "Has requirements.txt", "Has setup.py", "Requirements Packages Count",
        "Python Imports Count", "Package List", "Imports List"
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        cell.border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
    
    # Add branch data
    row_num = 2
    for repo in python_repos:
        repo_name = repo['name']
        
        # Sort branches to put default branch first
        branches = sorted(repo.get('branches', []), key=lambda b: (0 if b.get('is_default') else 1, b.get('name')))
        
        for branch in branches:
            # Extract packages from requirements.txt
            packages = branch.get('requirements_packages', [])
            package_list = ", ".join([f"{p['name']}{p['version'] if p['version'] != 'latest' else ''}" for p in packages])
            
            # Extract imports from Python files
            imports = branch.get('python_imports', [])
            # Filter out standard library modules for the display
            non_std_imports = [imp for imp, is_std in imports if not is_std]
            imports_list = ", ".join(sorted(non_std_imports))
            
            # Color coding
            if branch.get('is_default'):
                # Use light yellow for default branch
                fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            else:
                # Use light gray for other branches
                fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
            
            # Add branch data
            ws.cell(row=row_num, column=1).value = repo_name
            ws.cell(row=row_num, column=2).value = branch.get('name')
            ws.cell(row=row_num, column=3).value = "Yes" if branch.get('is_default') else "No"
            ws.cell(row=row_num, column=4).value = "Yes" if branch.get('has_pyproject') else "No"
            ws.cell(row=row_num, column=5).value = "Yes" if branch.get('has_requirements') else "No"
            ws.cell(row=row_num, column=6).value = "Yes" if branch.get('has_setup_py') else "No"
            ws.cell(row=row_num, column=7).value = len(packages)
            ws.cell(row=row_num, column=8).value = len(non_std_imports)
            ws.cell(row=row_num, column=9).value = package_list
            ws.cell(row=row_num, column=10).value = imports_list
            
            # Add coloring for default branch
            if branch.get('is_default'):
                for col_num in range(1, 11):
                    ws.cell(row=row_num, column=col_num).fill = fill
            
            row_num += 1
    
    # Set column widths
    column_widths = {
        1: 25,  # Repository name
        2: 20,  # Branch name
        3: 10,  # Is Default
        4: 15,  # Has pyproject.toml
        5: 15,  # Has requirements.txt
        6: 15,  # Has setup.py
        7: 20,  # Requirements Packages Count
        8: 20,  # Python Imports Count
        9: 60,  # Package List
        10: 60,  # Imports List
    }
    
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Freeze the header row and repository column
    ws.freeze_panes = "B2"

def create_packages_matrix(workbook, python_repos):
    """Create a matrix of packages across repositories and branches"""
    ws = workbook.create_sheet(title="Packages Matrix")
    
    # Get a list of all unique packages across all repositories and branches
    all_packages = set()
    for repo in python_repos:
        for branch in repo.get('branches', []):
            for pkg in branch.get('requirements_packages', []):
                all_packages.add(pkg['name'])
    
    # Sort packages alphabetically
    all_packages = sorted(list(all_packages))
    
    # Add header row with package names
    ws.cell(row=1, column=1).value = "Repository"
    ws.cell(row=1, column=2).value = "Branch"
    
    for col_num, package in enumerate(all_packages, 3):
        ws.cell(row=1, column=col_num).value = package
        ws.cell(row=1, column=col_num).font = Font(bold=True)
        ws.cell(row=1, column=col_num).alignment = Alignment(textRotation=90, horizontal='center')
    
    # Add repository and branch rows
    row_num = 2
    for repo in python_repos:
        repo_name = repo['name']
        
        # Sort branches to put default branch first
        branches = sorted(repo.get('branches', []), key=lambda b: (0 if b.get('is_default') else 1, b.get('name')))
        
        for branch in branches:
            branch_name = branch.get('name')
            
            # First two columns: repository and branch names
            ws.cell(row=row_num, column=1).value = repo_name
            ws.cell(row=row_num, column=2).value = branch_name
            
            # Set styling for default branch
            if branch.get('is_default'):
                ws.cell(row=row_num, column=1).font = Font(bold=True)
                ws.cell(row=row_num, column=2).font = Font(bold=True)
                for col in range(1, 3):
                    ws.cell(row=row_num, column=col).fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            
            # Get packages in this branch
            branch_packages = {pkg['name']: pkg['version'] for pkg in branch.get('requirements_packages', [])}
            
            # Fill in the matrix
            for col_num, package in enumerate(all_packages, 3):
                if package in branch_packages:
                    version = branch_packages[package]
                    cell = ws.cell(row=row_num, column=col_num)
                    
                    # Display version if available
                    if version != 'latest':
                        cell.value = version
                    else:
                        cell.value = "✓"
                    
                    # Highlight cell
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
            
            row_num += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25  # Repository name
    ws.column_dimensions['B'].width = 20  # Branch name
    
    # Set width for package columns
    for col_num in range(3, len(all_packages) + 3):
        ws.column_dimensions[get_column_letter(col_num)].width = 8
    
    # Freeze first two columns and header row
    ws.freeze_panes = "C2"

def create_imports_sheet(workbook, python_repos):
    """Create a sheet showing imports from Python files"""
    ws = workbook.create_sheet(title="Python Imports")
    
    # Get a list of all unique non-standard library imports across all repos and branches
    all_imports = set()
    for repo in python_repos:
        for branch in repo.get('branches', []):
            # Filter out standard library imports
            non_std_imports = [imp for imp, is_std in branch.get('python_imports', []) if not is_std]
            all_imports.update(non_std_imports)
    
    # Sort imports alphabetically
    all_imports = sorted(list(all_imports))
    
    # Add header row with import names
    ws.cell(row=1, column=1).value = "Repository"
    ws.cell(row=1, column=2).value = "Branch"
    
    for col_num, imp in enumerate(all_imports, 3):
        ws.cell(row=1, column=col_num).value = imp
        ws.cell(row=1, column=col_num).font = Font(bold=True)
        ws.cell(row=1, column=col_num).alignment = Alignment(textRotation=90, horizontal='center')
    
    # Add repository and branch rows
    row_num = 2
    for repo in python_repos:
        repo_name = repo['name']
        
        # Sort branches to put default branch first
        branches = sorted(repo.get('branches', []), key=lambda b: (0 if b.get('is_default') else 1, b.get('name')))
        
        for branch in branches:
            branch_name = branch.get('name')
            
            # First two columns: repository and branch names
            ws.cell(row=row_num, column=1).value = repo_name
            ws.cell(row=row_num, column=2).value = branch_name
            
            # Set styling for default branch
            if branch.get('is_default'):
                ws.cell(row=row_num, column=1).font = Font(bold=True)
                ws.cell(row=row_num, column=2).font = Font(bold=True)
                for col in range(1, 3):
                    ws.cell(row=row_num, column=col).fill = PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid")
            
            # Get imports in this branch
            branch_imports = {imp for imp, is_std in branch.get('python_imports', []) if not is_std}
            
            # Fill in the matrix
            for col_num, imp in enumerate(all_imports, 3):
                if imp in branch_imports:
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = "✓"
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    cell.alignment = Alignment(horizontal='center')
            
            row_num += 1
    
    # Set column widths
    ws.column_dimensions['A'].width = 25  # Repository name
    ws.column_dimensions['B'].width = 20  # Branch name
    
    # Set width for import columns
    for col_num in range(3, len(all_imports) + 3):
        ws.column_dimensions[get_column_letter(col_num)].width = 10
    
    # Freeze first two columns and header row
    ws.freeze_panes = "C2"

def create_summary_sheet(workbook, python_repos):
    """Create a summary sheet with key statistics"""
    ws = workbook.create_sheet(title="Summary")
    
    # Calculate statistics
    total_repos = len(python_repos)
    
    # Count branches
    all_branches = sum(len(repo.get('branches', [])) for repo in python_repos)
    default_branches_with_python = sum(1 for repo in python_repos 
                                      for branch in repo.get('branches', []) 
                                      if branch.get('is_default'))
    
    # Count packages from requirements files
    package_counts = {}
    for repo in python_repos:
        for branch in repo.get('branches', []):
            for pkg in branch.get('requirements_packages', []):
                package_counts[pkg['name']] = package_counts.get(pkg['name'], 0) + 1
    
    # Count non-standard library imports from Python files
    import_counts = {}
    for repo in python_repos:
        for branch in repo.get('branches', []):
            for imp, is_std in branch.get('python_imports', []):
                if not is_std:  # Skip standard library imports
                    import_counts[imp] = import_counts.get(imp, 0) + 1
    
    # Get most common packages from requirements
    top_packages = sorted(package_counts.items(), key=lambda x: x[1], reverse=True)[:20]
    
    # Get most common imports from Python files
    top_imports = sorted(import_counts.items(), key=lambda x: x[1], reverse=True)[:20]
    
    # Add summary data
    row = 1
    ws.cell(row=row, column=1).value = "Python Repositories Analysis"
    ws.cell(row=row, column=1).font = Font(bold=True, size=14)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    
    row += 2
    ws.cell(row=row, column=1).value = "Total Repositories with Python Files:"
    ws.cell(row=row, column=2).value = total_repos
    
    row += 1
    ws.cell(row=row, column=1).value = "Total Branches with Python Files:"
    ws.cell(row=row, column=2).value = all_branches
    
    row += 1
    ws.cell(row=row, column=1).value = "Default Branches with Python Files:"
    ws.cell(row=row, column=2).value = default_branches_with_python
    ws.cell(row=row, column=3).value = f"{default_branches_with_python/total_repos*100:.1f}%" if total_repos > 0 else "0%"
    
    row += 1
    ws.cell(row=row, column=1).value = "Total Unique Packages (from requirements):"
    ws.cell(row=row, column=2).value = len(package_counts)
    
    row += 1
    ws.cell(row=row, column=1).value = "Total Unique Imports (from Python files):"
    ws.cell(row=row, column=2).value = len(import_counts)
    
    # Add top packages section
    row += 2
    ws.cell(row=row, column=1).value = "Top 20 Most Common Packages (from requirements.txt)"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    
    row += 1
    ws.cell(row=row, column=1).value = "Package"
    ws.cell(row=row, column=2).value = "Count"
    ws.cell(row=row, column=3).value = "Percentage of Branches"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3).font = Font(bold=True)
    
    for package, count in top_packages:
        row += 1
        ws.cell(row=row, column=1).value = package
        ws.cell(row=row, column=2).value = count
        ws.cell(row=row, column=3).value = f"{count/all_branches*100:.1f}%" if all_branches > 0 else "0%"
    
    # Add top imports section
    row += 3  # Add some space
    ws.cell(row=row, column=1).value = "Top 20 Most Common Imports (from Python files)"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    
    row += 1
    ws.cell(row=row, column=1).value = "Import"
    ws.cell(row=row, column=2).value = "Count"
    ws.cell(row=row, column=3).value = "Percentage of Branches"
    ws.cell(row=row, column=1).font = Font(bold=True)
    ws.cell(row=row, column=2).font = Font(bold=True)
    ws.cell(row=row, column=3).font = Font(bold=True)
    
    for imp, count in top_imports:
        row += 1
        ws.cell(row=row, column=1).value = imp
        ws.cell(row=row, column=2).value = count
        ws.cell(row=row, column=3).value = f"{count/all_branches*100:.1f}%" if all_branches > 0 else "0%"
    
    # Set column widths
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 20

def main():
    """Main function to execute the script"""
    print("Starting GitHub repository and Python file analyzer...")
    
    # Check if the token has the necessary permissions
    token = os.environ.get('GITHUB_TOKEN')
    if token:
        try:
            headers = {'Authorization': f'token {token}'}
            scopes_response = requests.get('https://api.github.com/rate_limit', headers=headers)
            
            if 'X-OAuth-Scopes' in scopes_response.headers:
                scopes = scopes_response.headers['X-OAuth-Scopes'].split(', ')
                print("Token has the following scopes:", scopes)
                
                # Check for necessary scopes
                missing_scopes = []
                if 'repo' not in scopes:
                    missing_scopes.append('repo')
                if not any(s in scopes for s in ['read:org', 'admin:org', 'org']):
                    missing_scopes.append('read:org')
                
                if missing_scopes:
                    print("\nWARNING: Your token is missing important permissions:")
                    for scope in missing_scopes:
                        print(f"- Missing '{scope}' scope, which is needed to access all repositories")
                    print("\nTo create a new token with proper permissions:")
                    print("1. Go to https://github.com/settings/tokens")
                    print("2. Click 'Generate new token'")
                    print("3. Add the following scopes: repo, read:org")
                    print("4. Set the new token in your environment variables")
                    print("\nContinuing with current token, but results may be limited...")
                    print("---------------------------------------------------------------")
        except Exception as e:
            print(f"Error checking token permissions: {e}")
    
    # Get limited repositories in the organization (just 10 for testing)
    repos_limit = 183  # Limit to first 10 repositories
    print(f"Running with a limit of {repos_limit} repositories for testing purposes")
    repos = get_all_org_repos(limit=repos_limit)
    
    if not repos:
        print("No repositories found or an error occurred.")
        return
    
    # Find repositories with Python project files and analyze Python imports
    python_repos = find_python_project_files(repos)
    
    if not python_repos:
        print("No Python projects found in the organization repositories.")
        return
    
    print(f"\nFound {len(python_repos)} repositories with Python files.")
    
    # Create Excel report with all the collected information
    create_excel_report(python_repos)
    
    print("\nScript completed successfully.")
    print(f"Note: This run was limited to analyzing {repos_limit} repositories for testing.")
    print("To scan all repositories, update the limit in the main() function.")

if __name__ == "__main__":
    main()